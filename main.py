"""from fastapi import FastAPI, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

app = FastAPI()

@app.get("/")
def root():
    return {"status": "ok"}

@app.post("/generate_catalog")
async def generate_catalog(excel: UploadFile = File(...), images: List[UploadFile] = File(...)):
    wb = load_workbook(filename=io.BytesIO(await excel.read()))
    ws = wb.active

    for idx, image_file in enumerate(images, start=2):
        img_data = await image_file.read()
        image = Image.open(io.BytesIO(img_data))
        output = io.BytesIO()
        image.save(output, format='PNG')
        output.seek(0)
        img = XLImage(output)
        img.width = 100
        img.height = 100
        ws.add_image(img, f"B{idx}")

    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return StreamingResponse(
        output_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalog_ready.xlsx"}
    )

from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # або конкретно: ["https://for-vercel-sigma.vercel.app"]
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image
import openpyxl

app = FastAPI()

# CORS для фронтенду
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # або вкажи домен Vercel: ["https://for-vercel-sigma.vercel.app"]
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def root():
    return {"status": "ok"}

@app.post("/generate_catalog")
async def generate_catalog(
    excel_file: UploadFile = File(...),
    images: List[UploadFile] = File(...)
):
    try:
        wb = load_workbook(filename=BytesIO(await excel_file.read()))
        ws = wb.active

        image_size = (150, 150)
        start_row = 2  # припускаємо, що заголовки — у 1-му рядку
        img_column = "A"  # колонка для вставки зображень

        for i, upload in enumerate(images):
            img_bytes = await upload.read()
            image = Image.open(BytesIO(img_bytes))
            image.thumbnail(image_size)

            img_io = BytesIO()
            image.save(img_io, format="PNG")
            img_io.seek(0)

            img = openpyxl.drawing.image.Image(img_io)
            cell = f"{img_column}{start_row + i}"
            ws.add_image(img, cell)
            ws.row_dimensions[start_row + i].height = 75

        for col in ws.columns:
            for cell in col:
                if cell.column_letter == 'B':  # припустимо, що опис у колонці B
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
            "Content-Disposition": "attachment; filename=generated_catalog.xlsx"
        })

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
        """

from fastapi import FastAPI, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

app = FastAPI()

# Дозвіл CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # або конкретно: ["https://your-vercel-domain.vercel.app"]
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def generate_filename(base: str = "catalog", ext: str = ".xlsx") -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_{timestamp}{ext}"

def resize_image_high_quality(image_bytes: bytes, size=(150, 150)) -> BytesIO:
    img = Image.open(BytesIO(image_bytes)).convert("RGB")
    img.thumbnail(size, Image.LANCZOS)
    output = BytesIO()
    img.save(output, format="JPEG", quality=95, optimize=True)
    output.seek(0)
    return output

@app.get("/")
def root():
    return {"status": "ok"}

@app.post("/generate_catalog")
async def generate_catalog(
    excel_file: UploadFile = File(...),
    images: List[UploadFile] = File(...)
):
    try:
        wb = load_workbook(filename=BytesIO(await excel_file.read()))
        ws = wb.active

        start_row = 2
        image_column = "A"
        row_height = 75

        for i, upload in enumerate(images):
            raw_image = await upload.read()
            img_io = resize_image_high_quality(raw_image)
            img = XLImage(img_io)
            img.width = 100
            img.height = 100
            cell = f"{image_column}{start_row + i}"
            ws.add_image(img, cell)
            ws.row_dimensions[start_row + i].height = row_height

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = generate_filename()
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
